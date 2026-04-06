"""Excel workbook writer for trade_hunter output."""

from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd

_FILLED_STAR = "\u2605"  # ★

# Only columns whose names differ between the DataFrame and the workbook header.
_COLUMN_RENAMES = {
    "Symbol": "Ticker",
    "Last Price": "Price",
    "IV Idx": "IVx",
    "IV Rank": "IVR",
    "IV %tile": "IVP",
}

_OUTPUT_COLUMNS = [
    "Ticker",
    "Sector Bucket",
    "Sector",
    "Option Type",
    "Expiration Date",
    "Earnings Date",
    "DTE",
    "Price",
    "Strike",
    "Bid",
    "Ask",
    "Spread%",
    "Delta",
    "Open Interest",
    "Trade Score",
    "Quant Rating",
    "Liquidity",
    "Growth",
    "Momentum",
    "IVx",
    "IVR",
    "IVP",
    "BPR",
]

_NUMBER_FORMATS: dict[str, str] = {
    "Expiration Date": "MM/DD/YYYY",
    "Earnings Date": "MM/DD/YYYY",
    "DTE": "0",
    "Strike": "0",
    "Open Interest": "0",
    "Price": "0.00",
    "Trade Score": "0.00",
    "Quant Rating": "0.00",
    "Bid": "0.00",
    "Ask": "0.00",
    "Delta": "0.00",
    "IVR": "0.00",
    "Spread%": "0.0%",
    "IVx": "0.0%",
    "IVP": "0.0%",
    "BPR": "$#,##0",
}


def _prepare_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Transform, rename, and reorder columns for workbook output.

    Returns a new DataFrame with exactly the 23 output columns in order.
    """
    out = df.copy()

    if out.empty:
        return pd.DataFrame(columns=_OUTPUT_COLUMNS)

    out = out.sort_values("Trade Score", ascending=False).reset_index(drop=True)

    # Compute Spread% as a fraction before renaming
    bid = out["Bid"].astype(float)
    ask = out["Ask"].astype(float)
    out["Spread%"] = (ask - bid) / ((ask + bid) / 2)

    # Normalize IV Idx and IV %tile to fractions if stored as whole-number percentages
    # (IV Idx arrives as a plain float after loader normalization, e.g. 36.7 not 0.367)
    for col in ("IV Idx", "IV %tile"):
        vals = pd.to_numeric(out[col], errors="coerce")
        if not out.empty and vals.max(skipna=True) > 1.0:
            out[col] = vals / 100.0
        else:
            out[col] = vals

    # Title-case Option Type
    out["Option Type"] = out["Option Type"].str.title()

    # Parse ISO date strings to datetime.date for openpyxl date formatting
    for col in ("Expiration Date", "Earnings Date"):
        out[col] = out[col].apply(lambda v: date.fromisoformat(str(v)[:10]))

    # Convert raw star string to "X stars" display text
    out["Liquidity"] = out["Liquidity"].apply(lambda v: f"{str(v).count(_FILLED_STAR)} stars")

    # Rename columns that differ between DataFrame and workbook header
    out = out.rename(columns=_COLUMN_RENAMES)

    return out[_OUTPUT_COLUMNS]


def write_workbook(
    bull_scored: pd.DataFrame,
    bear_scored: pd.DataFrame,
    output_dir: Path,
) -> Path:
    """Write BULL-ish and BEAR-ish worksheets to output_dir/trade_signals.xlsx.

    Each sheet is sorted descending by Trade Score.
    Creates output_dir if it does not exist.
    Overwrites any existing trade_signals.xlsx.

    Returns the Path to the written file.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "trade_signals.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default blank sheet

    for sheet_name, scored_df in (("BULL-ish", bull_scored), ("BEAR-ish", bear_scored)):
        prepared = _prepare_sheet(scored_df)
        ws = wb.create_sheet(title=sheet_name)

        # Write header row
        for col_idx, header in enumerate(prepared.columns, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data rows with per-cell number formats
        for row_idx, row_tuple in enumerate(prepared.itertuples(index=False), start=2):
            for col_idx, (header, value) in enumerate(zip(prepared.columns, row_tuple), start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                fmt = _NUMBER_FORMATS.get(header)
                if fmt:
                    cell.number_format = fmt

    wb.save(out_path)
    return out_path
