from pathlib import Path

import openpyxl

WORKSHEET_NAME = "daJournal"


def load_journal(
    worksheets_dir: Path,
    explicit_path: Path | None = None,
) -> tuple[frozenset[str], list[str]]:
    """Load the daJournal worksheet and return a frozenset of deduplicated active symbols.

    If explicit_path is provided it is used directly; otherwise the file is resolved
    as worksheets_dir / "journal.xlsx".

    The file is opened read-only via openpyxl. No modifications are ever made to the
    source file.

    Raises:
        FileNotFoundError: if the resolved path does not exist.
        ValueError: if worksheet 'daJournal' is not present in the workbook.
        ValueError: if the 'Symbol' column is missing from the worksheet.
    """
    warnings: list[str] = []

    path = explicit_path if explicit_path is not None else worksheets_dir / "journal.xlsx"

    if not path.exists():
        raise FileNotFoundError(f"Journal file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if WORKSHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"Worksheet '{WORKSHEET_NAME}' not found in {path.name}. "
                f"Available sheets: {wb.sheetnames}"
            )

        ws = wb[WORKSHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise ValueError(f"Worksheet '{WORKSHEET_NAME}' is empty")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if "Symbol" not in headers:
        raise ValueError(f"'Symbol' column not found in worksheet '{WORKSHEET_NAME}'")

    symbol_idx = headers.index("Symbol")

    null_count = 0
    raw_symbols: list[str] = []
    for row in rows[1:]:
        val = row[symbol_idx] if len(row) > symbol_idx else None
        if val is None or str(val).strip() == "":
            null_count += 1
        else:
            raw_symbols.append(str(val).strip())

    if null_count:
        warnings.append(f"[Journal] {null_count} row(s) with null/empty Symbol dropped")

    result = frozenset(raw_symbols)

    if not result:
        warnings.append("[Journal] No active trade symbols found after cleaning")

    return result, warnings
