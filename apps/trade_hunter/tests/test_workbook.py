"""Unit tests for write_workbook()."""

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from trade_hunter.output.workbook import _OUTPUT_COLUMNS, write_workbook
from trade_hunter.pipeline.scoring import SCORE_COLUMNS

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_FILLED_STAR = "\u2605"  # ★
_EMPTY_STAR = "\u2606"  # ☆


def _minimal_row(**overrides) -> dict:
    """Return a minimal valid scored DataFrame row."""
    row = {
        "Symbol": "AAPL",
        "Sector Bucket": "Growth",
        "Sector": "Information Technology",
        "Option Type": "put",
        "Expiration Date": "2025-04-18",
        "Earnings Date": "2025-05-28",
        "DTE": 30,
        "Last Price": 100.0,
        "Strike": 95.0,
        "Bid": 1.10,
        "Ask": 1.20,
        "Delta": -0.21,
        "Open Interest": 500,
        "Trade Score": 3.87,
        **{col: 3.0 for col in SCORE_COLUMNS},
        "Quant Rating": 4.0,
        "Liquidity": _FILLED_STAR * 3 + _EMPTY_STAR,  # ★★★☆
        "Growth": "B",
        "Momentum": "A-",
        "IV Idx": 22.5,
        "IV Rank": 45.0,
        "IV %tile": 55.0,
        "BPR": 1610.0,
    }
    row.update(overrides)
    return row


def _df(*rows: dict) -> pd.DataFrame:
    return pd.DataFrame(list(rows))


def _empty_scored() -> pd.DataFrame:
    return pd.DataFrame(columns=list(_minimal_row().keys()))


def _read_sheet(path: Path, sheet_name: str):
    """Return (headers, data_rows) from an xlsx sheet."""
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0]) if rows else []
    data = rows[1:] if len(rows) > 1 else []
    return headers, data


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_workbook_creates_file(tmp_path):
    bull = _df(_minimal_row())
    bear = _df(_minimal_row(Symbol="MSFT", Trade_Score=2.5))
    path = write_workbook(bull, bear, tmp_path)
    assert path == tmp_path / "trade_signals.xlsx"
    assert path.exists()


def test_workbook_sheet_names(tmp_path):
    path = write_workbook(_df(_minimal_row()), _df(_minimal_row()), tmp_path)
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["BULL-ish", "BEAR-ish"]


def test_workbook_column_headers(tmp_path):
    path = write_workbook(_df(_minimal_row()), _df(_minimal_row()), tmp_path)
    for sheet_name in ("BULL-ish", "BEAR-ish"):
        headers, _ = _read_sheet(path, sheet_name)
        assert headers == _OUTPUT_COLUMNS


def test_workbook_sorted_by_trade_score(tmp_path):
    rows = [
        _minimal_row(Symbol="A", **{"Trade Score": 3.0}),
        _minimal_row(Symbol="B", **{"Trade Score": 4.5}),
        _minimal_row(Symbol="C", **{"Trade Score": 2.1}),
    ]
    path = write_workbook(_df(*rows), _df(_minimal_row()), tmp_path)
    headers, data = _read_sheet(path, "BULL-ish")
    score_col = headers.index("Trade Score")
    first_score = data[0][score_col]
    assert first_score == 4.5


def test_workbook_overwrites_existing(tmp_path):
    bull_first = _df(_minimal_row(Symbol="FIRST", **{"Trade Score": 1.0}))
    bull_second = _df(_minimal_row(Symbol="SECOND", **{"Trade Score": 2.0}))
    bear = _df(_minimal_row())
    write_workbook(bull_first, bear, tmp_path)
    write_workbook(bull_second, bear, tmp_path)
    headers, data = _read_sheet(tmp_path / "trade_signals.xlsx", "BULL-ish")
    ticker_col = headers.index("Ticker")
    assert data[0][ticker_col] == "SECOND"


def test_workbook_creates_output_dir(tmp_path):
    subdir = tmp_path / "nested" / "output"
    assert not subdir.exists()
    write_workbook(_df(_minimal_row()), _df(_minimal_row()), subdir)
    assert subdir.exists()
    assert (subdir / "trade_signals.xlsx").exists()


def test_workbook_option_type_titlecased(tmp_path):
    path = write_workbook(
        _df(_minimal_row(**{"Option Type": "put"})), _df(_minimal_row()), tmp_path
    )
    headers, data = _read_sheet(path, "BULL-ish")
    col = headers.index("Option Type")
    assert data[0][col] == "Put"


def test_workbook_empty_bear_sheet(tmp_path):
    path = write_workbook(_df(_minimal_row()), _empty_scored(), tmp_path)
    wb = openpyxl.load_workbook(path)
    assert "BEAR-ish" in wb.sheetnames
    ws = wb["BEAR-ish"]
    rows = list(ws.iter_rows(values_only=True))
    assert len(rows) == 1  # header row only
    assert list(rows[0]) == _OUTPUT_COLUMNS


def test_workbook_liquidity_stars_text(tmp_path):
    liquidity = _FILLED_STAR * 3 + _EMPTY_STAR  # ★★★☆
    path = write_workbook(
        _df(_minimal_row(**{"Liquidity": liquidity})), _df(_minimal_row()), tmp_path
    )
    headers, data = _read_sheet(path, "BULL-ish")
    col = headers.index("Liquidity")
    assert data[0][col] == "3 stars"


def test_workbook_has_36_columns(tmp_path):
    """Output spreadsheet has 36 columns (23 original + 13 individual score columns)."""
    path = write_workbook(_df(_minimal_row()), _df(_minimal_row()), tmp_path)
    headers, _ = _read_sheet(path, "BULL-ish")
    assert len(headers) == 36


def test_workbook_score_columns_present_and_after_trade_score(tmp_path):
    """All 13 individual score columns appear immediately after Trade Score."""
    path = write_workbook(_df(_minimal_row()), _df(_minimal_row()), tmp_path)
    headers, _ = _read_sheet(path, "BULL-ish")
    trade_score_idx = headers.index("Trade Score")
    score_headers = headers[trade_score_idx + 1: trade_score_idx + 1 + len(SCORE_COLUMNS)]
    assert score_headers == SCORE_COLUMNS


def test_workbook_score_column_values(tmp_path):
    """Individual score column values are written correctly from the DataFrame."""
    row = _minimal_row(**{"IVR Score": 4.0, "Momentum Score": 2.5})
    path = write_workbook(_df(row), _df(_minimal_row()), tmp_path)
    headers, data = _read_sheet(path, "BULL-ish")
    assert data[0][headers.index("IVR Score")] == pytest.approx(4.0)
    assert data[0][headers.index("Momentum Score")] == pytest.approx(2.5)
