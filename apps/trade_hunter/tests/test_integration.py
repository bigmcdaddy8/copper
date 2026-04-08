"""End-to-end integration test for the full trade_hunter pipeline.

Uses synthetic input files written to tmp_path and a MagicMock TradierClient —
no real files or API calls are made.
"""

from datetime import date
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest

from trade_hunter.config import RunConfig
from trade_hunter.pipeline.runner import run_pipeline

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_RUN_DATE = date(2025, 3, 19)
# April 18, 2025 is the third Friday of April — DTE=30 from _RUN_DATE
_EXPIRATION = "2025-04-18"
_FILLED_STAR = "\u2605"  # ★
_EMPTY_STAR = "\u2606"  # ☆


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def pipeline_result(tmp_path):
    """Run the full pipeline with synthetic files and a mock Tradier client.

    Returns (workbook_path, log_path, output_dir).
    """
    downloads = tmp_path / "downloads"
    worksheets = tmp_path / "worksheets"
    output = tmp_path / "output"
    downloads.mkdir()
    worksheets.mkdir()

    # --- TastyTrade CSV ---
    tt_path = downloads / "tastytrade.csv"
    pd.DataFrame(
        {
            "Symbol": ["AAPL"],
            "Sector": ["Technology"],
            "Liquidity": [_FILLED_STAR * 3 + _EMPTY_STAR],  # ★★★☆
            "IV Idx": [22.5],
            "IV Rank": [45.0],
            "IV %tile": [55.0],
            "Earnings At": ["2025-05-28"],
            "Last": [100.0],
            "Name": ["Apple Inc."],
        }
    ).to_csv(tt_path, index=False)

    # --- SeekingAlpha BULL ---
    bull_path = downloads / "bull.xlsx"
    pd.DataFrame(
        {
            "Symbol": ["AAPL"],
            "Quant Rating": [4.0],
            "Growth": ["B"],
            "Momentum": ["A-"],
        }
    ).to_excel(bull_path, index=False, engine="openpyxl")

    # --- SeekingAlpha BEAR ---
    bear_path = downloads / "bear.xlsx"
    pd.DataFrame(
        {
            "Symbol": ["AAPL"],
            "Quant Rating": [4.0],
            "Growth": ["B"],
            "Momentum": ["A-"],
        }
    ).to_excel(bear_path, index=False, engine="openpyxl")

    # --- journal.xlsx with daJournal sheet ---
    journal_path = worksheets / "journal.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "daJournal"
    ws.append(["Symbol"])
    ws.append(["MSFT"])  # active trade — AAPL is not excluded
    wb.save(journal_path)

    # --- Mock TradierClient ---
    mock_client = MagicMock()
    mock_client.get_option_expirations.return_value = [_EXPIRATION]
    mock_client.get_last_price.return_value = 100.0
    mock_client.get_option_chain.return_value = [
        {
            "option_type": "put",
            "strike": 95.0,
            "delta": -0.21,
            "open_interest": 500,
            "bid": 1.10,
            "ask": 1.20,
        },
        {
            "option_type": "call",
            "strike": 105.0,
            "delta": 0.21,
            "open_interest": 300,
            "bid": 0.90,
            "ask": 1.00,
        },
    ]

    config = RunConfig(
        output_dir=output,
        tradier_api_key="test-key",
        tastytrade_file=tt_path,
        bull_file=bull_path,
        bear_file=bear_path,
        journal_file=journal_path,
        cache_dir=None,  # memory-only; avoid real yfinance calls
    )

    # Patch yfinance so sector lookups return a known value without network calls
    mock_yf_ticker = MagicMock()
    mock_yf_ticker.info = {"sector": "Technology"}
    with patch("trade_hunter.loaders.sector_cache.yfinance") as mock_yf:
        mock_yf.Ticker.return_value = mock_yf_ticker
        workbook_path, log_path = run_pipeline(config, mock_client, run_date=_RUN_DATE)
    return workbook_path, log_path, output


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _read_sheet(path: Path, sheet_name: str):
    wb = openpyxl.load_workbook(path)
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    headers = list(rows[0]) if rows else []
    data = rows[1:] if len(rows) > 1 else []
    return headers, data


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_integration_workbook_exists(pipeline_result):
    workbook_path, _, _ = pipeline_result
    assert workbook_path.exists()


def test_integration_log_exists(pipeline_result):
    _, log_path, _ = pipeline_result
    assert log_path.exists()


def test_integration_bull_sheet_has_data(pipeline_result):
    workbook_path, _, _ = pipeline_result
    _, data = _read_sheet(workbook_path, "BULL-ish")
    assert len(data) >= 1


def test_integration_bear_sheet_has_data(pipeline_result):
    workbook_path, _, _ = pipeline_result
    _, data = _read_sheet(workbook_path, "BEAR-ish")
    assert len(data) >= 1


def test_integration_bull_ticker_is_aapl(pipeline_result):
    workbook_path, _, _ = pipeline_result
    headers, data = _read_sheet(workbook_path, "BULL-ish")
    ticker_col = headers.index("Ticker")
    assert data[0][ticker_col] == "AAPL"


def test_integration_column_count(pipeline_result):
    workbook_path, _, _ = pipeline_result
    for sheet_name in ("BULL-ish", "BEAR-ish"):
        headers, _ = _read_sheet(workbook_path, sheet_name)
        assert len(headers) == 36, f"{sheet_name} has {len(headers)} columns, expected 36"
